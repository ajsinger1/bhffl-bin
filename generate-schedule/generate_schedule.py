"""
generate_schedule.py

random fantasy football schedule generator

backtracking algorithm using MRV and forward checking heuristic
"""

import csv
import random
import argparse
from itertools import combinations, product, permutations
from copy import deepcopy
from collections import defaultdict
from tqdm import tqdm
import multiprocessing
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NUM_WEEKS = 13
teams = None


class Multiset:
    def __init__(self):
        self.dict = {}

    def __init__(self, iterable):
        self.dict = {}
        for el in iterable:
            self.add(el)

    def __len__(self):
        return sum(self.dict.values())

    def __contains__(self, el):
        return el in self.dict

    def __iter__(self):
        for el in self.dict.keys():
            for _ in range(self.dict[el]):
                yield el

    def add(self, el):
        if el in self.dict:
            self.dict[el] += 1
        else:
            self.dict[el] = 1

    def remove(self, el):
        if el in self.dict:
            self.dict[el] -= 1
            if self.dict[el] == 0:
                del self.dict[el]


class GameGraph:
    def __init__(self, all_games, teams):
        self.teams = set(teams)
        self.graph = defaultdict(set)
        self.weeks = [set() for _ in range(NUM_WEEKS)]

        for game in all_games:
            self._add_game(game)
        assert all(
            [len(edges) == (len(teams) - 2) * 2 + 1 for edges in self.graph.values()]
        )

    def schedule_week(self, game_to_schedule, week_idx):
        assert week_idx >= 0 and week_idx < NUM_WEEKS
        if self._can_schedule_week(game_to_schedule, week_idx):
            assert game_to_schedule not in self.weeks[week_idx]
            self.weeks[week_idx].add(game_to_schedule)
            return True
        return False

    def deschedule_week(self, game_to_deschedule, week_idx):
        assert week_idx >= 0 and week_idx < NUM_WEEKS
        assert game_to_deschedule in self.weeks[week_idx]
        self.weeks[week_idx].remove(game_to_deschedule)

    def get_games_for_week(self, week_idx):
        return len(self.weeks[week_idx])

    def get_num_schedulable_weeks(self, game_to_schedule):
        num_schedulable_weeks = 0
        for week in self.weeks:
            if not any(
                [
                    game_to_schedule[0] in game or game_to_schedule[1] in game
                    for game in week
                ]
            ):
                num_schedulable_weeks += 1
        return num_schedulable_weeks

    def get_num_scheduled(self):
        return sum([len(week) for week in self.weeks])

    def validate_graph(self):
        for matchups in self.weeks:
            assert len(matchups) == 6
            for game1, game2 in combinations(matchups, 2):
                assert game2 not in self.graph[game1]
                assert game1 not in self.graph[game2]

    def print(self):
        for week, matchups in enumerate(self.weeks, 1):
            print(f"Week {week}:")
            for game in matchups:
                print()
                print(f"  {game[0]} vs {game[1]}")
            print("-" * 15)

    def save(self, file):
        if file.endswith(".csv"):
            self._save_csv(file)
            return
        if not file.endswith(".xlsx"):
            file += ".xlsx"
        self._save_xlsx(file)

    def _can_schedule_week(self, game_to_schedule, week_idx):
        # Week is full
        if len(self.weeks[week_idx]) >= len(self.teams) // 2:
            return False

        # Conflicting game
        if any([game in self.graph[game_to_schedule] for game in self.weeks[week_idx]]):
            return False

        return True

    def _add_game(self, new_game):
        new_edges = set()
        for game in self.graph.keys():
            if game[0] in new_game or game[1] in new_game:
                new_edges.add(game)
                self.graph[game].add(new_game)
        new_edges.add(new_game)
        self.graph[new_game] = new_edges

    def _save_csv(self, file):
        with open(file, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Week", "Home Team", "Away Team"])
            for week, games in enumerate(self.weeks, 1):
                for game in games:
                    writer.writerow([week, game[0], game[1]])
            print(f"Schedule has been written to {file}")

    def _save_xlsx(self, file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # Styles
        week_header_font = Font(name="Cordia New", bold=True, size=22, color="FFFFFF")
        col_header_font = Font(name="Cordia New", bold=True, size=18, color="FFFFFF")
        cell_font = Font(name="Cordia New", bold=True, size=16, color="000000")
        header_fill = PatternFill(
            start_color="7B3EAC", end_color="7B3EAC", fill_type="solid"
        )
        centered = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Grid layout parameters
        columns_per_grid = 3
        rows_between_grids = 2
        cols_between_grids = 1

        # Column widths
        team_col_width = 15
        vs_col_width = 5

        # Write schedule
        for week, games in enumerate(self.weeks, 1):
            # Calculate grid position
            grid_row = (week - 1) // columns_per_grid
            grid_col = (week - 1) % columns_per_grid

            start_row = grid_row * (len(games) + 3 + rows_between_grids) + 1
            start_col = grid_col * (3 + cols_between_grids) + 1

            # Week header
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=start_row,
                end_column=start_col + 2,
            )
            week_cell = ws.cell(row=start_row, column=start_col, value=f"Week {week}")
            week_cell.font = week_header_font
            week_cell.fill = header_fill
            week_cell.alignment = centered
            for col in range(start_col, start_col + 3):
                ws.cell(row=start_row, column=col).border = border

            # Column headers
            headers = ["Home Team", "vs", "Away Team"]
            for col, header in enumerate(headers, start_col):
                cell = ws.cell(row=start_row + 1, column=col, value=header)
                cell.font = col_header_font
                cell.fill = header_fill
                cell.alignment = centered
                cell.border = border

            # Games
            for i, game in enumerate(games):
                row = start_row + 2 + i
                cells = [
                    ws.cell(row=row, column=start_col, value=game[0]),
                    ws.cell(row=row, column=start_col + 1, value="vs"),
                    ws.cell(row=row, column=start_col + 2, value=game[1]),
                ]
                for cell in cells:
                    cell.border = border
                    cell.alignment = centered
                    cell.font = cell_font

        # Adjust column widths
        for col in range(1, ws.max_column + 1):
            if (col - 1) % (3 + cols_between_grids) == 1:  # 'vs' columns
                ws.column_dimensions[get_column_letter(col)].width = vs_col_width
            else:  # team name columns
                ws.column_dimensions[get_column_letter(col)].width = team_col_width

        wb.save(file)
        print(f"Schedule has been written to {file}")


def get_most_constrained_game(game_graph, remaining_games):
    least_weeks_possible = min(
        [game_graph.get_num_schedulable_weeks(game) for game in remaining_games]
    )
    return random.choice(
        list(
            filter(
                lambda x: game_graph.get_num_schedulable_weeks(x)
                == least_weeks_possible,
                remaining_games,
            )
        )
    )


def generate_schedule_util(game_graph, remaining_games, bar):
    if len(remaining_games) == 0:
        game_graph.validate_graph()
        return True

    # Choose game which is most constrained first
    game_to_schedule = get_most_constrained_game(game_graph, remaining_games)
    remaining_games.remove(game_to_schedule)
    bar.update(1)

    # Try scheduling weeks with the most amount of games first
    weeks_sorted = sorted(
        [week_idx for week_idx in range(NUM_WEEKS)],
        reverse=True,
        key=game_graph.get_games_for_week,
    )
    for week in weeks_sorted:
        if game_graph.schedule_week(game_to_schedule, week):
            assert game_graph.get_num_scheduled() == 78 - len(remaining_games)
            if generate_schedule_util(game_graph, remaining_games, bar):
                return True
            else:
                game_graph.deschedule_week(game_to_schedule, week)
    # At this point we have tried all weeks and none work, need to backtrack
    bar.update(-1)
    remaining_games.add(game_to_schedule)
    return False


def read_teams(file_path):
    teams = {f"DIVISION {i+1}": [] for i in range(4)}
    with open(file_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            for division, team in row.items():
                if team:
                    teams[division].append(team)
    return teams


def generate_schedule(teams):
    all_teams = [team for division in teams.values() for team in division]

    # Generate in-division games (each team plays others in division twice)
    in_division_games = []
    for division in teams.values():
        in_division_games.extend(list(combinations(division, 2)) * 2)

    # Generate out-of-division games (each team plays all others once)
    out_of_division_games = []
    for div1, div2 in combinations(teams.values(), 2):
        out_of_division_games.extend(product(div1, div2))

    all_games = Multiset(in_division_games + out_of_division_games)

    game_graph = GameGraph(all_games, all_teams)

    bar = tqdm(total=len(all_games))

    assert generate_schedule_util(game_graph, all_games, bar)
    return game_graph


def generate_schedule_wrapper(teams, return_dict):
    try:
        schedule = generate_schedule(teams)
        return_dict["result"] = schedule
    except Exception as e:
        return_dict["error"] = str(e)


def generate_schedule_with_timeout(teams, timeout=5):
    manager = multiprocessing.Manager()
    return_dict = manager.dict()
    process = multiprocessing.Process(
        target=generate_schedule_wrapper, args=(teams, return_dict)
    )

    process.start()
    process.join(timeout)

    if process.is_alive():
        process.terminate()
        process.join()
        return None

    if "result" in return_dict:
        return return_dict["result"]
    elif "error" in return_dict:
        raise Exception(return_dict["error"])
    else:
        return None


def main():
    parser = argparse.ArgumentParser(
        description="Generate a fantasy football schedule from a CSV file."
    )
    parser.add_argument(
        "-f",
        "--file",
        help="Path to the input CSV file containing team names",
        default="example_divisions.csv",
    )
    parser.add_argument(
        "-i", "--info", action="store_true", help="Display information about the script"
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Path to output CSV/XSLX file for the schedule",
        default="schedule",
    )
    parser.add_argument(
        "-n",
        "--no-output",
        help="Do not save output file",
        action="store_true",
        default=False,
    )
    args = parser.parse_args()

    if args.info:
        print("Fantasy Football Schedule Generator")
        print("===================================")
        print(
            "This script generates a 13-week fantasy football schedule based on the teams provided in a CSV file."
        )
        print(
            "The CSV file should have four columns: DIVISION 1, DIVISION 2, DIVISION 3, DIVISION 4."
        )
        print("Each column should contain three team names.")
        print("\nThe generated schedule ensures that:")
        print("- Every team plays the other two teams in their division twice")
        print("- Every team plays all teams from other divisions once")
        print("- The schedule is randomized each time the script is run")
        return

    teams = read_teams(args.file)
    schedule = None
    for attempt in range(1, 11):
        # Sometimes the algorithm gets "unlucky" and ends up running for 1+ minutes
        # Most of the time it generates a schedule in <5s, so we will terminate generation and retry at 5s
        print(f"Attempt {attempt} to generate schedule...")
        schedule = generate_schedule_with_timeout(teams)
        if schedule:
            print(f"Schedule generated successfully after {attempt} attempt(s).")
            break
        else:
            print(f"Attempt {attempt} failed. Retrying...")

    if not schedule:
        print(f"Failed to generate schedule after {args.max_attempts} attempts.")
        exit(1)

    schedule.print()
    if args.output and not args.no_output:
        schedule.save(args.output)


if __name__ == "__main__":
    main()
